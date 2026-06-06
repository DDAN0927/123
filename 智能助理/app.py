# -*- coding: utf-8 -*-
import streamlit as st
import os
import base64
import io
import uuid
import threading
from dotenv import load_dotenv
from memory import get_vectordb
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeout

load_dotenv()

st.set_page_config(page_title="SmartPal Agent", page_icon="🧠", layout="wide")

INVALID_KEY_PATTERNS = [
    "sk-your-", "your-", "你的", "替换", "placeholder", "here", "xxx", "test",
]

IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"}
TEXT_EXTENSIONS = {".txt", ".md", ".py", ".js", ".html", ".css", ".json", ".xml", ".yaml", ".yml", ".csv", ".log"}
DOCUMENT_EXTENSIONS = {".pdf", ".docx", ".xlsx", ".xls"}

_agent_lock = threading.Lock()


def _is_valid_api_key(key: str) -> tuple:
    if not key or not key.strip():
        return False, "API Key 为空，请在左侧边栏填写"
    if any(p in key.lower() for p in INVALID_KEY_PATTERNS):
        return False, "API Key 无效（包含占位符），请填写真实的 API Key"
    try:
        key.encode("ascii")
    except UnicodeEncodeError:
        return False, "API Key 包含非英文字符，请检查是否填写了正确的 Key"
    if len(key.strip()) < 10:
        return False, "API Key 长度太短，请检查是否填写完整"
    return True, ""


def _extract_text_from_pdf(bytes_data: bytes) -> str:
    from PyPDF2 import PdfReader
    reader = PdfReader(io.BytesIO(bytes_data))
    texts = []
    for page in reader.pages:
        t = page.extract_text()
        if t:
            texts.append(t)
    return "\n".join(texts)


def _extract_text_from_docx(bytes_data: bytes) -> str:
    from docx import Document
    doc = Document(io.BytesIO(bytes_data))
    return "\n".join([p.text for p in doc.paragraphs if p.text.strip()])


def _extract_text_from_xlsx(bytes_data: bytes) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(bytes_data), read_only=True)
    texts = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            line = " | ".join([str(c) for c in row if c is not None])
            if line.strip():
                texts.append(line)
    wb.close()
    return "\n".join(texts)


def _process_uploaded_file(uploaded_file) -> dict:
    name = uploaded_file.name.lower()
    ext = os.path.splitext(name)[1]
    bytes_data = uploaded_file.read()

    if ext in IMAGE_EXTENSIONS:
        b64 = base64.b64encode(bytes_data).decode("utf-8")
        mime_map = {
            ".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
            ".gif": "image/gif", ".webp": "image/webp", ".bmp": "image/bmp",
        }
        mime = mime_map.get(ext, "image/png")
        return {
            "type": "image",
            "filename": uploaded_file.name,
            "base64": b64,
            "mime": mime,
            "data_url": f"data:{mime};base64,{b64}",
        }

    text_content = ""
    if ext in TEXT_EXTENSIONS:
        for enc in ["utf-8", "gbk", "gb2312", "latin-1"]:
            try:
                text_content = bytes_data.decode(enc)
                break
            except (UnicodeDecodeError, LookupError):
                continue
    elif ext == ".pdf":
        text_content = _extract_text_from_pdf(bytes_data)
    elif ext == ".docx":
        text_content = _extract_text_from_docx(bytes_data)
    elif ext in {".xlsx", ".xls"}:
        text_content = _extract_text_from_xlsx(bytes_data)
    else:
        for enc in ["utf-8", "gbk", "latin-1"]:
            try:
                text_content = bytes_data.decode(enc)
                break
            except (UnicodeDecodeError, LookupError):
                continue

    return {
        "type": "file",
        "filename": uploaded_file.name,
        "content": text_content[:8000],
    }


with st.sidebar:
    st.title("⚙️ 配置")

    api_key = st.text_input(
        "API Key",
        value=os.getenv("OPENAI_API_KEY", ""),
        type="password",
        help="阿里云 DashScope / DeepSeek / OpenAI 兼容 API Key"
    )
    base_url = st.text_input(
        "API 地址",
        value=os.getenv("OPENAI_BASE_URL", "https://dashscope.aliyuncs.com/compatible-mode/v1"),
        help="API Base URL，阿里云 DashScope / DeepSeek / OpenAI 代理"
    )
    model = st.selectbox(
        "模型",
        ["qwen-turbo", "qwen-plus", "qwen-max", "qwen-vl-plus", "qwen-vl-max", "deepseek-chat", "deepseek-reasoner", "gpt-4o-mini"],
        index=0,
        help="选择模型。图片识别请选 qwen-vl-plus 或 qwen-vl-max"
    )

    valid, key_msg = _is_valid_api_key(api_key)
    if not valid:
        st.warning(f"⚠️ {key_msg}")

    if st.button("🔄 应用配置"):
        valid, key_msg = _is_valid_api_key(api_key)
        if not valid:
            st.error(f"❌ {key_msg}")
        else:
            os.environ["OPENAI_API_KEY"] = api_key.strip()
            os.environ["OPENAI_BASE_URL"] = base_url.strip()
            os.environ["OPENAI_MODEL"] = model
            st.session_state.messages = []
            st.session_state.agent = None
            st.success("✅ 配置已更新！")

    if st.button("🔍 测试连接"):
        valid, key_msg = _is_valid_api_key(api_key)
        if not valid:
            st.error(f"❌ {key_msg}")
        else:
            import httpx
            target_url = base_url if base_url else "https://api.openai.com/v1"
            endpoint = f"{target_url.rstrip('/')}/models"
            headers = {"Authorization": f"Bearer {api_key.strip()}"}
            with st.spinner("正在测试连接..."):
                try:
                    client = httpx.Client(timeout=httpx.Timeout(connect=8.0, read=10.0))
                    resp = client.get(endpoint, headers=headers)
                    if resp.status_code == 200:
                        st.success(f"✅ 连接成功！({target_url})")
                    elif resp.status_code == 401:
                        st.error("❌ API Key 无效 (401)，请检查 Key 是否正确")
                    elif resp.status_code == 403:
                        st.error("❌ 无权限访问 (403)，请检查 Key 或账户状态")
                    elif resp.status_code == 402:
                        st.error("❌ 余额不足 (402)，请充值后重试")
                    else:
                        st.warning(f"⚠️ 返回状态 {resp.status_code}: {resp.text[:200]}")
                except httpx.ConnectTimeout:
                    st.error(f"❌ 连接超时，无法连接到 {target_url}")
                except httpx.ConnectError:
                    st.error(f"❌ 无法连接，{target_url} 不可达")
                except UnicodeEncodeError:
                    st.error("❌ API Key 包含非英文字符，请检查 Key 是否正确")
                except Exception as ex:
                    st.error(f"❌ 连接错误: {str(ex)[:200]}")

    st.divider()
    st.title("🧠 思考模式")
    enable_reflection = st.toggle(
        "深度思考",
        value=True,
        help="开启后，SmartPal 会在回答前自我反思，给出更有深度的回答"
    )
    show_thinking = st.toggle(
        "展示思考过程",
        value=True,
        help="开启后，你可以看到 SmartPal 的思考链"
    )

    st.divider()
    st.caption(f"API: {base_url}")
    st.caption(f"Model: {model}")

st.title("🧠 SmartPal — 有想法的智能助理")

if "messages" not in st.session_state:
    st.session_state.messages = []
if "agent" not in st.session_state:
    st.session_state.agent = None

for msg in st.session_state.messages:
    with st.chat_message(msg["role"]):
        if msg.get("images"):
            for img in msg["images"]:
                st.image(img, width=300)
        if msg.get("files"):
            for fn in msg["files"]:
                st.caption(f"📎 {fn}")
        st.markdown(msg["content"])
        if show_thinking and msg.get("thinking"):
            with st.expander("💭 思考过程", expanded=False):
                for step in msg["thinking"]:
                    st.markdown(f"> {step}")


def _retrieve_context(prompt: str) -> str:
    try:
        db = get_vectordb()
        if db:
            docs = db.similarity_search(prompt, k=2)
            return "\n".join([d.page_content for d in docs])
    except Exception:
        pass
    return ""


def get_or_create_agent():
    if st.session_state.agent is not None:
        return st.session_state.agent

    current_key = api_key.strip()
    current_url = base_url.strip()
    current_model = model

    valid, key_msg = _is_valid_api_key(current_key)
    if not valid:
        st.error(f"❌ {key_msg}，请在左侧边栏填写正确的 API Key 后点击「应用配置」")
        return None

    try:
        import httpx
        from langchain_openai import ChatOpenAI

        from tools import (
            search_web, get_weather, get_current_time,
            calculate, write_note, read_note, add_reminder, list_reminders
        )
        from react_agent import ReActAgent

        TOOLS = [
            search_web,
            get_weather,
            get_current_time,
            calculate,
            write_note,
            read_note,
            add_reminder,
            list_reminders,
        ]

        if not current_url and "deepseek" in current_model.lower():
            current_url = "https://api.deepseek.com"
        elif not current_url and "qwen" in current_model.lower():
            current_url = "https://dashscope.aliyuncs.com/compatible-mode/v1"

        http_timeout = httpx.Timeout(connect=10.0, read=90.0, write=10.0, pool=10.0)
        http_client = httpx.Client(timeout=http_timeout)

        llm_kwargs = {
            "model": current_model,
            "temperature": 0,
            "openai_api_key": current_key,
            "max_retries": 2,
            "http_client": http_client,
        }
        if current_url:
            llm_kwargs["openai_api_base"] = current_url

        llm = ChatOpenAI(**llm_kwargs)

        SYSTEM_PROMPT = """你是 SmartPal，一个有独立思考能力的智能助手。

## 你的性格
- 你有自己的想法和观点，不是简单的问答机器
- 你会主动思考用户真正需要什么，而不只是回答表面问题
- 你会在回答后主动给出相关建议或延伸思考
- 你有好奇心，遇到有趣的话题会分享你的见解
- 你的语气友好但有自己的态度，像一个聪明的朋友

## 思考框架
每次收到问题，按以下框架思考：

1. **理解**：用户真正想问什么？有没有隐含需求？
2. **分析**：这个问题需要什么信息？我是否需要工具辅助？
3. **思考**：我对这个问题有什么看法？有没有更深入的视角？
4. **行动**：如果需要工具，调用工具获取信息
5. **反思**：我的回答是否完整？有没有遗漏？用户可能还想知道什么？
6. **回答**：给出有深度的回答，包含你的思考和建议

## 工具使用格式
```
Thought: 你的思考过程（包含理解、分析、你的观点）
Action: 工具名称
Action Input: {"参数名": "参数值"}
```

如果不需要工具，直接回答：
```
Thought: 你的思考过程
Reflection: 自我审视——我的回答是否完整？有没有遗漏？用户可能还想知道什么？
Answer: 你的回答（包含你的见解和主动建议）
```

## 工具使用指南
- 用户问时间 → get_current_time
- 用户问天气 → get_weather（查完天气后主动建议穿衣/出行）
- 用户需要计算 → calculate
- 用户要记笔记 → write_note
- 用户要搜索 → search_web
- 搜索/查天气后，主动给出你的分析和建议

## 回答原则
- 不只回答问题，还要给出你的思考和见解
- 主动建议相关的后续行动或信息
- 如果发现用户可能有未表达的隐含需求，主动提出
- 回答要有深度，但不要啰嗦
- 用户上传图片时，仔细观察图片内容并给出你的分析和见解
- 用户上传文件时，基于文件内容回答问题并给出你的建议
"""

        agent = ReActAgent(
            llm=llm,
            tools=TOOLS,
            system_prompt=SYSTEM_PROMPT,
            max_iterations=5,
            enable_reflection=enable_reflection,
        )
        st.session_state.agent = agent
        return agent
    except UnicodeEncodeError:
        st.error("❌ API Key 包含非英文字符，请检查 Key 是否正确填写")
        return None
    except Exception as e:
        st.error(f"❌ Agent 初始化失败: {e}")
        return None


uploaded_files = st.file_uploader(
    "📎 上传图片或文件",
    type=["png", "jpg", "jpeg", "gif", "webp", "bmp", "txt", "md", "py", "js",
          "html", "css", "json", "xml", "yaml", "yml", "csv", "log", "pdf", "docx", "xlsx", "xls"],
    accept_multiple_files=True,
    key="file_uploader",
)

if prompt := st.chat_input("输入你的需求..."):
    user_images = []
    user_filenames = []
    file_context_parts = []

    if uploaded_files:
        for f in uploaded_files:
            f.seek(0)
            processed = _process_uploaded_file(f)
            if processed["type"] == "image":
                user_images.append(processed["data_url"])
                user_filenames.append(processed["filename"])
            else:
                user_filenames.append(processed["filename"])
                if processed.get("content"):
                    file_context_parts.append(
                        f"--- 文件: {processed['filename']} ---\n{processed['content']}\n"
                    )

    display_content = prompt
    if user_filenames:
        display_content = f"📎 {', '.join(user_filenames)}\n\n{prompt}"

    st.session_state.messages.append({
        "id": str(uuid.uuid4()),
        "role": "user",
        "content": display_content,
        "images": user_images,
        "files": user_filenames,
        "file_context": file_context_parts,
        "responded": False,
    })
    st.rerun()

unanswered = [m for m in st.session_state.messages if m["role"] == "user" and not m.get("responded")]
if unanswered:
    msg = unanswered[-1]
    acquired = _agent_lock.acquire(blocking=False)
    if not acquired:
        st.info("⏳ 正在思考中，请稍候...")
    else:
        msg["responded"] = True
        try:
            agent = get_or_create_agent()
            if agent is None:
                st.session_state.messages.append({
                    "role": "assistant",
                    "content": "❌ Agent 未初始化，请检查左侧边栏的 API Key 配置。",
                })
            else:
                prompt_text = msg["content"]
                if msg.get("files"):
                    for fn in msg["files"]:
                        prompt_text = prompt_text.replace(f"📎 {fn}", "").strip()
                user_images = msg.get("images", [])
                file_context_parts = msg.get("file_context", [])

                context = ""
                try:
                    with ThreadPoolExecutor(max_workers=1) as pool:
                        future = pool.submit(_retrieve_context, prompt_text)
                        try:
                            context = future.result(timeout=10)
                        except FutureTimeout:
                            context = ""
                except Exception:
                    context = ""

                full_input = ""
                if context:
                    full_input += f"已知信息：{context}\n\n"
                if file_context_parts:
                    full_input += "用户上传的文件内容：\n" + "\n".join(file_context_parts) + "\n\n"
                full_input += f"用户问题：{prompt_text}"

                response = ""
                thinking = []
                try:
                    with st.spinner("🧠 思考中..."):
                        invoke_input = {"input": full_input}
                        if user_images:
                            invoke_input["images"] = user_images
                        result = agent.invoke(invoke_input)
                        response = result["output"]
                        thinking = result.get("thinking", [])
                except UnicodeEncodeError:
                    response = "❌ **API Key 格式错误** — API Key 包含非英文字符，请检查配置。"
                except ConnectionError:
                    response = "❌ **网络连接失败** — 无法连接到 API 服务器，请检查网络和 API 地址。"
                except Exception as e:
                    error_msg = str(e)
                    if "401" in error_msg or "AuthenticationError" in type(e).__name__:
                        response = "❌ **API Key 无效 (401)** — 请检查 API Key 是否正确，然后在左侧边栏更新后点击「应用配置」。"
                        st.session_state.agent = None
                    elif "Connection" in error_msg or "connect" in error_msg.lower() or "timeout" in error_msg.lower():
                        response = f"❌ **网络连接失败**\n\n原始错误: `{error_msg}`"
                    elif "429" in error_msg:
                        response = f"❌ **请求过于频繁 (429)**\n\n原始错误: `{error_msg}`"
                    elif "402" in error_msg:
                        response = f"❌ **余额不足 (402)**\n\n原始错误: `{error_msg}`"
                    else:
                        response = f"❌ **请求失败**\n\n原始错误: `{error_msg}`"

                st.session_state.messages.append({
                    "role": "assistant",
                    "content": response,
                    "thinking": thinking,
                })
        finally:
            _agent_lock.release()
        st.rerun()
